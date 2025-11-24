import os
import re
import sqlite3
from datetime import datetime
from io import BytesIO

from flask import (
    Flask, render_template, request, redirect,
    url_for, flash, send_file, session
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "parking_secret_key"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "parking.db")

LOGIN_PASSWORD = "1112"


# -----------------------------
# DB 관련
# -----------------------------
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            plate TEXT NOT NULL,
            memo TEXT,
            UNIQUE(date, plate)
        );
        """
    )
    conn.commit()
    conn.close()


# Flask 3.x: before_first_request 대신 플래그 사용 + 로그인 체크
db_initialized = False


@app.before_request
def before_request():
    global db_initialized
    if not db_initialized:
        init_db()
        db_initialized = True

    # 로그인 안 해도 되는 엔드포인트
    allowed = {"login", "static"}
    if request.endpoint in allowed or request.endpoint is None:
        return

    if not session.get("logged_in"):
        return redirect(url_for("login"))


# -----------------------------
# 차량번호 처리
# 형식: 2~3숫자 + 한글 + 공백 + 4숫자
# 예) 11저 8604 , 145마 9820
# -----------------------------
PLATE_CORE_REGEX = re.compile(r"^([0-9]{2,3}[가-힣])([0-9]{4})$")


def normalize_plate(raw: str):
    """
    - 공백 제거
    - 2~3숫자 + 한글 + 4숫자 검사
    - '앞부분 뒷부분' 형식으로 리턴 (예: '11저 8604')
    """
    if not raw:
        return None

    txt = re.sub(r"\s+", "", raw)

    m = PLATE_CORE_REGEX.match(txt)
    if not m:
        return None

    front, back = m.group(1), m.group(2)
    return f"{front} {back}"


# -----------------------------
# 로그인
# -----------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        pw = (request.form.get("password") or "").strip()
        if pw == LOGIN_PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("index"))
        else:
            flash("비밀번호가 올바르지 않습니다.")

    return render_template("login.html")


# -----------------------------
# 메인 화면
# -----------------------------
@app.route("/", methods=["GET"])
def index():
    conn = get_conn()
    cur = conn.cursor()

    today = datetime.now().strftime("%Y-%m-%d") # 등록 기본 날짜
    view_date = request.args.get("view_date") or today

    # 날짜 포맷 검증 (이상하면 오늘로)
    try:
        datetime.strptime(view_date, "%Y-%m-%d")
    except ValueError:
        view_date = today

    query_plate = request.args.get("q", "").strip()

    # 선택 날짜 주차 차량 목록 + 누적횟수
    cur.execute(
        "SELECT id, plate FROM logs WHERE date=? ORDER BY time, plate",
        (view_date,),
    )
    day_logs = cur.fetchall()

    day_rows = []
    for row in day_logs:
        plate = row["plate"]
        log_id = row["id"]
        cur.execute("SELECT COUNT(*) AS cnt FROM logs WHERE plate=?", (plate,))
        total_cnt = cur.fetchone()["cnt"]
        day_rows.append({"id": log_id, "plate": plate, "total_cnt": total_cnt})

    day_count = len(day_rows)

    # 전체 차량번호 (자동완성용)
    cur.execute("SELECT DISTINCT plate FROM logs ORDER BY plate")
    all_plates = [row["plate"] for row in cur.fetchall()]

    # 차량 조회 (누적 + 최초입차일 + 최근입차일)
    search_result = None
    if query_plate:
        norm = normalize_plate(query_plate)
        if not norm:
            flash("차량번호 형식이 올바르지 않습니다. 예) 11저 8604")
        else:
            cur.execute(
                """
                SELECT
                    COUNT(*) AS cnt,
                    MIN(date) AS first_date,
                    MAX(date) AS last_date
                FROM logs
                WHERE plate=?
                """,
                (norm,),
            )
            row = cur.fetchone()
            if row["cnt"] > 0:
                search_result = {
                    "plate": norm,
                    "total_cnt": row["cnt"],
                    "first_date": row["first_date"],
                    "last_date": row["last_date"],
                }
            else:
                search_result = {
                    "plate": norm,
                    "total_cnt": 0,
                    "first_date": None,
                    "last_date": None,
                }

    conn.close()

    return render_template(
        "index.html",
        today=today,
        view_date=view_date,
        day_rows=day_rows,
        day_count=day_count,
        all_plates=all_plates,
        query_plate=query_plate,
        search_result=search_result,
    )


# -----------------------------
# 신규 주차 차량 여러 대 추가
# -----------------------------
@app.route("/add", methods=["POST"])
def add():
    date = request.form.get("date") or datetime.now().strftime("%Y-%m-%d")
    plates = request.form.getlist("plate")

    now_time = datetime.now().strftime("%H:%M")

    conn = get_conn()
    cur = conn.cursor()

    added = 0
    duplicated = 0
    invalid = 0

    for plate_input in plates:
        plate_input = (plate_input or "").strip()
        if not plate_input:
            continue

        plate_norm = normalize_plate(plate_input)
        if not plate_norm:
            invalid += 1
            continue

        try:
            cur.execute(
                "INSERT INTO logs(date, time, plate, memo) VALUES (?,?,?,?)",
                (date, now_time, plate_norm, ""),
            )
            added += 1
        except sqlite3.IntegrityError:
            duplicated += 1

    conn.commit()
    conn.close()

    msg_parts = []
    if added:
        msg_parts.append(f"저장 {added}건")
    if duplicated:
        msg_parts.append(f"중복 {duplicated}건")
    if invalid:
        msg_parts.append(f"형식오류 {invalid}건")
    if not msg_parts:
        msg_parts.append("입력된 차량이 없습니다.")

    flash(" / ".join(msg_parts))
    return redirect(url_for("index"))


# -----------------------------
# 선택 날짜 기록 삭제 (X 버튼)
# -----------------------------
@app.route("/delete", methods=["POST"])
def delete():
    log_id = request.form.get("id")
    if not log_id:
        flash("삭제할 항목을 찾지 못했습니다.")
        return redirect(url_for("index"))

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM logs WHERE id=?", (log_id,))
    affected = cur.rowcount
    conn.commit()
    conn.close()

    if affected:
        flash("삭제되었습니다.")
    else:
        flash("이미 삭제되었거나 존재하지 않는 항목입니다.")

    return redirect(url_for("index"))


# -----------------------------
# DB 백업 (parking.db 통째로 다운로드)
# -----------------------------
@app.route("/backup")
def backup():
    if not os.path.exists(DB_PATH):
        flash("백업할 데이터가 없습니다.")
        return redirect(url_for("index"))

    today_str = datetime.now().strftime("%Y-%m-%d")
    return send_file(
        DB_PATH,
        as_attachment=True,
        mimetype="application/octet-stream",
        download_name=f"parking_backup_{today_str}.db",
    )


# -----------------------------
# 엑셀(xlsx) 내보내기
# (차량번호 / 누적주차횟수 만)
# -----------------------------
def build_excel(rows, title: str):
    """
    rows: (plate, total_cnt) 튜플 리스트
    title: 시트 상단 제목
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "주차기록"

    # 제목 (A1:B1 합침)
    ws.merge_cells("A1:B1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 헤더
    headers = ["차량번호", "누적주차횟수"]
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 데이터 (3행부터)
    row_idx = 3
    for (plate, total_cnt) in rows:
        c1 = ws.cell(row=row_idx, column=1, value=plate)
        c2 = ws.cell(row=row_idx, column=2, value=total_cnt)

        for cell in (c1, c2):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        row_idx += 1

    # 열 너비 (고정값으로 넉넉하게)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15

    # 헤더 아래로 고정
    ws.freeze_panes = "A3"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@app.route("/export")
def export():
    scope = request.args.get("scope", "day")
    conn = get_conn()
    cur = conn.cursor()

    if scope == "all":
        # 전체 기간: 차량번호별 누적 주차 횟수
        cur.execute(
            "SELECT plate, COUNT(*) AS cnt "
            "FROM logs "
            "GROUP BY plate "
            "ORDER BY plate"
        )
        rows = cur.fetchall()
        title = "주차 차량 누적 횟수 (전체)"
        filename = "parking_all.xlsx"
    else:
        # 특정 날짜 기준: 해당 날짜에 온 차량만, 누적 주차 횟수
        date_str = request.args.get("date")
        if not date_str:
            date_str = datetime.now().strftime("%Y-%m-%d")

        cur.execute(
            "SELECT plate, COUNT(*) AS cnt "
            "FROM logs "
            "WHERE date=? "
            "GROUP BY plate "
            "ORDER BY plate",
            (date_str,),
        )
        rows = cur.fetchall()
        title = f"주차 차량 누적 횟수 ({date_str} 기준)"
        filename = f"parking_{date_str}.xlsx"

    conn.close()

    data_rows = [(r["plate"], r["cnt"]) for r in rows]
    bio = build_excel(data_rows, title=title)

    return send_file(
        bio,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        as_attachment=True,
        download_name=filename,
    )


# -----------------------------
# 로컬 실행용
# -----------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)